from __future__ import annotations

import gc
import os
import re

import openpyxl
import pythoncom
import win32com.client
from openpyxl.utils import column_index_from_string

from .config import GRANTED_COL, PUB_COL, JUSTIA_BASE, MAX_CHARS, ROW_HEIGHT
from .debug import dlog

# XML 1.0 forbids most C0 control characters except \t \n \r
_INVALID_XML_CHARS = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")


# ──────────────────────────────────────────────────────────────────────────────
# Scan
# ──────────────────────────────────────────────────────────────────────────────

def scan_rows_to_process(
    filepath: str,
    start_row: int = 2,
    url_col: str = "AV",
    single_row: bool = False,
) -> list[tuple[int, str]]:
    """Return ALL (row_num, url) pairs found in the URL column.

    Whether to skip already-fetched rows is decided by the caller
    (ExcelProcessor) based on txt file presence, not Excel cell content.
    """
    dlog(f"scan_rows_to_process: filepath={filepath} start_row={start_row} "
         f"url_col={url_col} single_row={single_row}")

    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active
    col_idx = column_index_from_string(url_col.upper().strip())

    scan_start = max(start_row, 2)
    scan_end   = start_row if single_row else (ws.max_row or 1)
    dlog(f"scan_rows_to_process: col={col_idx} max_row={ws.max_row} "
         f"scan={scan_start}..{scan_end}")

    read_min = min(col_idx, GRANTED_COL, PUB_COL)
    read_max = max(col_idx, GRANTED_COL, PUB_COL)

    def cv(t, c):
        i = c - read_min
        return t[i] if 0 <= i < len(t) else None

    url_col_has_data = False
    url_rows:    list[tuple[int, str]] = []
    patent_rows: list[tuple[int, str]] = []

    for row_num, row_tuple in enumerate(
        ws.iter_rows(min_row=scan_start, max_row=scan_end,
                     min_col=read_min,   max_col=read_max,
                     values_only=True),
        start=scan_start,
    ):
        url_val = cv(row_tuple, col_idx)
        has_url = bool(url_val and str(url_val).strip()
                       and not str(url_val).startswith("="))

        if has_url:
            url_col_has_data = True
            url_rows.append((row_num, str(url_val).strip()))
        else:
            granted = str(cv(row_tuple, GRANTED_COL) or "").strip()
            pub     = str(cv(row_tuple, PUB_COL)     or "").strip()
            if granted or pub:
                patent_rows.append((row_num, JUSTIA_BASE + (granted or pub)))

    wb.close()

    if url_col_has_data:
        dlog(f"scan_rows_to_process: url_col 模式，{len(url_rows)} 列")
        return url_rows

    dlog(f"scan_rows_to_process: patent fallback，{len(patent_rows)} 列")
    return patent_rows


# ──────────────────────────────────────────────────────────────────────────────
# Text splitting
# ──────────────────────────────────────────────────────────────────────────────

_BREAK_CHARS = ('\n', '.', '。', ';', '；', '!', '！', '?', '？')

def split_text(text: str, max_chars: int = 32000) -> list[str]:
    """Split text into chunks ≤ max_chars.

    Break priority: newline > sentence-ending punctuation > hard cut.
    """
    chunks: list[str] = []
    while len(text) > max_chars:
        # 找每種斷點在 max_chars 以內的最後位置，取最靠右者
        cut = max(
            (text.rfind(ch, 0, max_chars) for ch in _BREAK_CHARS),
            default=-1,
        )
        if cut < 0:
            cut = max_chars - 1
        chunks.append(text[:cut + 1])
        text = text[cut + 1:]
    if text:
        chunks.append(text)
    return chunks


# ──────────────────────────────────────────────────────────────────────────────
# Excel COM writer
# ──────────────────────────────────────────────────────────────────────────────

class ExcelComWriter:
    """Writes cell values via Excel COM (preserves all existing styles/formulas)."""

    def __init__(self, filepath: str):
        self._filepath = os.path.abspath(filepath)
        dlog(f"ExcelComWriter.__init__: {self._filepath}")
        pythoncom.CoInitialize()
        self._excel = win32com.client.DispatchEx("Excel.Application")
        self._excel.Visible        = False
        self._excel.DisplayAlerts  = False
        self._excel.ScreenUpdating = False
        self._wb = self._excel.Workbooks.Open(self._filepath)
        self._ws = self._wb.Sheets(1)
        dlog("ExcelComWriter.__init__: 完成")

    def add_hyperlink(self, row_num: int, col: int, address: str, display: str = "開啟") -> None:
        # Positional args: Anchor, Address, SubAddress, ScreenTip, TextToDisplay
        self._ws.Hyperlinks.Add(
            self._ws.Cells(row_num, col),
            address,
            "",
            "",
            display,
        )
        dlog(f"ExcelComWriter.add_hyperlink: 第{row_num}列 col={col} → {address} [{display}]")

    def write_chunks(self, row_num: int, start_col: int, chunks: list[str]) -> None:
        """Write each chunk to consecutive columns; append an empty terminator column."""
        for i, chunk in enumerate(chunks):
            safe = _INVALID_XML_CHARS.sub("", chunk)
            self._ws.Cells(row_num, start_col + i).Value = safe
            self._ws.Rows(row_num).RowHeight = ROW_HEIGHT
            dlog(f"ExcelComWriter.write_chunks: 第{row_num}列 col={start_col+i} {len(safe)}字")
        # empty terminator
        self._ws.Cells(row_num, start_col + len(chunks)).Value = ""

    def save(self) -> None:
        self._wb.Save()
        dlog("ExcelComWriter.save: 完成")

    def close(self) -> None:
        try:
            self._wb.Close(SaveChanges=True)
            self._excel.Quit()
            dlog("ExcelComWriter.close: 完成")
        finally:
            del self._ws
            del self._wb
            del self._excel
            gc.collect()
            pythoncom.CoUninitialize()

    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.close()


# ──────────────────────────────────────────────────────────────────────────────
# Post-processing: txt files → xlsx
# ──────────────────────────────────────────────────────────────────────────────

def write_xlsx_from_txts(
    filepath: str,
    row_txt_pairs: list[tuple[int, str]],
    url_col_idx: int,
    status_cb=None,
    stop_event=None,
    log_cb=None,
) -> None:
    """Write char count at url_col+3, chunk count at url_col+4, content at url_col+5+."""
    if not row_txt_pairs:
        return

    hyperlink_col   = url_col_idx + 3
    char_count_col  = url_col_idx + 4
    chunk_count_col = url_col_idx + 5
    content_col     = url_col_idx + 6
    xlsx_dir        = os.path.dirname(os.path.abspath(filepath))

    dlog(f"write_xlsx_from_txts: {len(row_txt_pairs)} 列，"
         f"cols=+2(link)/+3(chars)/+4(chunks)/+5+(content)")
    if status_cb:
        status_cb(f"寫入 Excel（共 {len(row_txt_pairs)} 列）…")

    written = 0
    with ExcelComWriter(filepath) as writer:
        for i, (row_num, txt_path) in enumerate(row_txt_pairs, 1):
            if stop_event and stop_event.is_set():
                dlog("write_xlsx_from_txts: 收到停止訊號")
                writer.save()
                if status_cb:
                    status_cb(f"已停止（已寫入 {written} 列）")
                return

            try:
                with open(txt_path, encoding="utf-8") as f:
                    text = f.read()
            except OSError as e:
                dlog(f"write_xlsx_from_txts: 讀取 {txt_path} 失敗：{e}")
                text = ""

            chunks = split_text(text)

            # 1. 超連結（右2），顯示檔名含副檔名
            rel = os.path.relpath(txt_path, xlsx_dir)
            writer.add_hyperlink(row_num, hyperlink_col, rel,
                                 display=os.path.basename(txt_path))

            # 2. 寫入各 chunk（內容）
            writer.write_chunks(row_num, content_col, chunks)

            # 3. 回填實際占用格數
            cell_chunks = writer._ws.Cells(row_num, chunk_count_col)
            cell_chunks.Style = "Normal"
            cell_chunks.Value = str(len(chunks))

            # 4. 最後寫字元數
            cell_chars = writer._ws.Cells(row_num, char_count_col)
            cell_chars.Style = "Normal"
            cell_chars.Value = str(len(text))

            dlog(f"write_xlsx_from_txts: 第{row_num}列 → {len(text)}字 {len(chunks)}格")
            written += 1
            if log_cb:
                log_cb(f"第{row_num}列 → {len(text)}字 {len(chunks)}格\n")
            if status_cb and i % 10 == 0:
                status_cb(f"寫入 Excel… {i}/{len(row_txt_pairs)}")

        writer.save()

    if status_cb:
        status_cb("完成")
    dlog("write_xlsx_from_txts: 全部完成")
